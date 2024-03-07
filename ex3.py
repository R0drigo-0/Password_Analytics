import os
import nltk
import numpy as np
import openpyxl
from gensim.models import Word2Vec
from nltk.tokenize import word_tokenize
from gensim.test.utils import common_texts
from scipy.optimize import linear_sum_assignment
from sklearn.feature_extraction.text import TfidfVectorizer

same_user = {}

model = Word2Vec(
    sentences=common_texts, vector_size=100, window=5, min_count=1, workers=4
)
model.save("word2vec.model")


def get_word_embedding(word):
    """
    Returns the word embedding for a given word.
    """
    # Check if the word is in the model's vocabulary
    if word in model.wv:
        # Return the word's embedding
        return model.wv[word]
    else:
        # Handle out-of-vocabulary words (e.g., by returning a zero vector)
        return np.zeros(model.vector_size)


def calculate_distance_matrix(sentence1_embeddings, sentence2_embeddings):
    # Initialize an empty matrix to store the distances
    distance_matrix = np.zeros((len(sentence1_embeddings), len(sentence2_embeddings)))

    # Calculate the Euclidean distance between each pair of word embeddings
    for i, emb1 in enumerate(sentence1_embeddings):
        for j, emb2 in enumerate(sentence2_embeddings):
            distance_matrix[i, j] = np.linalg.norm(emb1 - emb2)

    return distance_matrix


def word_mover_distance(sentence1, sentence2):
    # Obtain word embeddings for each token in the sentences
    sentence1_embeddings = np.array(
        [get_word_embedding(sentence1)]
    )
    sentence2_embeddings = np.array(
        [get_word_embedding(sentence2)]
    )

    # Calculate the distance matrix between embeddings
    distance_matrix = calculate_distance_matrix(
        sentence1_embeddings, sentence2_embeddings
    )

    # Solve the transportation problem to find the optimal assignment
    row_ind, col_ind = linear_sum_assignment(distance_matrix)

    # Calculate the WMD as the sum of the distances for the optimal assignment
    wmd = np.sum(distance_matrix[row_ind, col_ind])

    return wmd


folder_path = "pass/"
files = os.listdir(folder_path)
for file in files:
    full_path = os.path.join(folder_path, file)
    with open(full_path, "r", encoding="iso-8859-1") as f:
        lines = f.readlines()

        for line in lines:
            correo = ""
            contraseña = ""

            if ":" in line:
                correo, contraseña = line.split(":", 1)
            elif ";" in line:
                correo, contraseña = line.split(";", 1)
            else:
                print("Password not found")
                break

            user = correo.split("@")[0]
            if user not in same_user:
                same_user[user] = []
            same_user[user].append(contraseña)

print("Processing data")
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write header
worksheet.cell(row=1, column=1).value = "Usuario"
worksheet.cell(row=1, column=2).value = "Contraseñas"

# Write data starting from row 2
row_index = 2
for user, contraseñas in same_user.items():
    related_passwords = []
    for i in range(len(contraseñas)):
        for j in range(i + 1, len(contraseñas)):
            distancia_wm = word_mover_distance(contraseñas[i], contraseñas[j])
            if distancia_wm <= 0.1:
                related_passwords.append(contraseñas[i])
                related_passwords.append(contraseñas[j])
    # Remove duplicates and keep the order of insertion
    related_passwords = list(dict.fromkeys(related_passwords))
    if related_passwords:
        worksheet.cell(row=row_index, column=1).value = user
        worksheet.cell(row=row_index, column=2).value = ", ".join(related_passwords)
        row_index += 1

# Save the workbook
workbook.save("pasword_by_user.xlsx")

print("All related passwords stored in 'pasword_by_user.xlsx'.")